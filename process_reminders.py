import os
import sys
import glob
import datetime
import openpyxl
import warnings
import ctypes
from ctypes import wintypes
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Suppress openpyxl UserWarnings
warnings.simplefilter("ignore", UserWarning)

# ==========================================
# Clipboard File Copy Support (Windows CF_HDROP)
# ==========================================
CF_HDROP = 15

class POINT(ctypes.Structure):
    _fields_ = [("x", wintypes.LONG), ("y", wintypes.LONG)]

class DROPFILES(ctypes.Structure):
    _fields_ = [
        ("pFiles", wintypes.DWORD),
        ("pt", POINT),
        ("fNC", wintypes.BOOL),
        ("fWide", wintypes.BOOL)
    ]

def copy_file_to_clipboard(file_path):
    """Copies a physical file to the Windows clipboard (CF_HDROP)."""
    if os.name != 'nt':
        return False
    try:
        # 显式声明 kernel32/user32 函数的参数和返回值类型以避免 64 位指针截断
        ctypes.windll.kernel32.GlobalAlloc.restype = wintypes.HANDLE
        ctypes.windll.kernel32.GlobalAlloc.argtypes = [wintypes.UINT, ctypes.c_size_t]
        
        ctypes.windll.kernel32.GlobalLock.restype = wintypes.LPVOID
        ctypes.windll.kernel32.GlobalLock.argtypes = [wintypes.HANDLE]
        
        ctypes.windll.kernel32.GlobalUnlock.restype = wintypes.BOOL
        ctypes.windll.kernel32.GlobalUnlock.argtypes = [wintypes.HANDLE]
        
        ctypes.windll.kernel32.GlobalFree.restype = wintypes.HANDLE
        ctypes.windll.kernel32.GlobalFree.argtypes = [wintypes.HANDLE]

        ctypes.windll.user32.OpenClipboard.restype = wintypes.BOOL
        ctypes.windll.user32.OpenClipboard.argtypes = [wintypes.HWND]

        ctypes.windll.user32.SetClipboardData.restype = wintypes.HANDLE
        ctypes.windll.user32.SetClipboardData.argtypes = [wintypes.UINT, wintypes.HANDLE]

        abs_path = os.path.abspath(file_path).replace('/', '\\')
        files_str = abs_path + "\0\0"
        files_bytes = files_str.encode('utf-16le')
        
        dropfiles_size = ctypes.sizeof(DROPFILES)
        GHND = 0x0042 # GMEM_MOVEABLE | GMEM_ZEROINIT
        
        hGlobal = ctypes.windll.kernel32.GlobalAlloc(GHND, dropfiles_size + len(files_bytes))
        if not hGlobal:
            print("【调试】GlobalAlloc 失败")
            return False
            
        pGlobal = ctypes.windll.kernel32.GlobalLock(hGlobal)
        if not pGlobal:
            print("【调试】GlobalLock 失败")
            ctypes.windll.kernel32.GlobalFree(hGlobal)
            return False
            
        try:
            df = DROPFILES()
            df.pFiles = dropfiles_size
            df.fWide = True
            
            ctypes.memmove(pGlobal, ctypes.byref(df), dropfiles_size)
            ctypes.memmove(pGlobal + dropfiles_size, files_bytes, len(files_bytes))
        finally:
            ctypes.windll.kernel32.GlobalUnlock(hGlobal)
            
        if ctypes.windll.user32.OpenClipboard(None):
            try:
                ctypes.windll.user32.EmptyClipboard()
                ctypes.windll.user32.SetClipboardData(CF_HDROP, hGlobal)
            finally:
                ctypes.windll.user32.CloseClipboard()
            return True
        else:
            err = ctypes.GetLastError()
            print(f"【调试】打开剪贴板失败，GetLastError(): {err} (通常是因为当前进程在非交互式后台服务中运行，本地双击运行即可成功)")
            ctypes.windll.kernel32.GlobalFree(hGlobal)
            return False
    except Exception as e:
        print(f"【警告】复制文件到剪贴板失败: {e}")
        return False

# ==========================================
# 1. 2026 Holiday & Workday Adjust Data
# ==========================================
HOLIDAYS = {
    # New Year
    "2026-01-01", "2026-01-02", "2026-01-03",
    # Spring Festival
    "2026-02-15", "2026-02-16", "2026-02-17", "2026-02-18", "2026-02-19",
    "2026-02-20", "2026-02-21", "2026-02-22", "2026-02-23",
    # Tomb Sweeping
    "2026-04-04", "2026-04-05", "2026-04-06",
    # Labor Day
    "2026-05-01", "2026-05-02", "2026-05-03", "2026-05-04", "2026-05-05",
    # Dragon Boat
    "2026-06-19", "2026-06-20", "2026-06-21",
    # Mid-Autumn
    "2026-09-25", "2026-09-26", "2026-09-27",
    # National Day
    "2026-10-01", "2026-10-02", "2026-10-03", "2026-10-04", "2026-10-05",
    "2026-10-06", "2026-10-07"
}

WORKDAY_ADJUSTS = {
    "2026-01-04",
    "2026-02-14", "2026-02-28",
    "2026-05-09",
    "2026-09-20", "2026-10-10"
}

def calculate_working_days_decimal(start_date, end_date):
    """Replicates VBA CalculateWorkingDaysDecimal calculation."""
    if not isinstance(start_date, datetime.datetime):
        if isinstance(start_date, datetime.date):
            start_date = datetime.datetime.combine(start_date, datetime.time.min)
        else:
            return 0.0
    if not isinstance(end_date, datetime.datetime):
        if isinstance(end_date, datetime.date):
            end_date = datetime.datetime.combine(end_date, datetime.time.min)
        else:
            return 0.0

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    # Strip times to get calendar days difference
    start_day = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_day = end_date.replace(hour=0, minute=0, second=0, microsecond=0)
    total_days = (end_day - start_day).days

    work_days = 0.0

    for i in range(total_days + 1):
        current_day = start_day + datetime.timedelta(days=i)
        date_str = current_day.strftime("%Y-%m-%d")

        is_holiday = date_str in HOLIDAYS
        is_workday_adjust = date_str in WORKDAY_ADJUSTS
        # weekday() in python: 0=Mon, 6=Sun. <= 4 means Mon-Fri.
        is_weekday = current_day.weekday() <= 4

        if not (is_workday_adjust or (not is_holiday and is_weekday)):
            continue

        day_start = current_day
        day_end = day_start + datetime.timedelta(days=1)

        if i == 0:
            # First day fraction (VBA: dayEnd - startDate)
            fraction = (day_end - start_date).total_seconds() / 86400.0
            work_days += fraction
        elif i == total_days:
            # Last day fraction (VBA: endDate - dayStart)
            fraction = (end_date - day_start).total_seconds() / 86400.0
            work_days += fraction
        else:
            # Full day
            work_days += 1.0

    # Return raw unrounded float (rounded to 6 decimal places to prevent float representation inaccuracy)
    return round(work_days, 6)


def is_last_month_received(change_time, stats_time):
    """Replicates VBA 判断是否上月领出."""
    if not isinstance(change_time, (datetime.datetime, datetime.date)) or not isinstance(stats_time, (datetime.datetime, datetime.date)):
        return "否"
    change_year, change_month = change_time.year, change_time.month
    stats_year, stats_month = stats_time.year, stats_time.month

    if change_year == stats_year:
        if change_month == stats_month - 1:
            return "是"
        elif change_month == stats_month:
            return "否"
        else:
            return "是（更早）"
    elif change_year == stats_year - 1:
        if change_month == 12 and stats_month == 1:
            return "是"
        else:
            return "是（更早）"
    else:
        return "是（更早）"


def simplify_district_name(name):
    """Replicates VBA SimplifyDistrictName."""
    if not name:
        return ""
    name_str = str(name).strip()
    if "城区" in name_str:
        return "城区局"
    elif "东北" in name_str:
        return "东北局"
    elif "东南" in name_str:
        return "东南局"
    elif "东区" in name_str:
        return "东区局"
    elif "虎门" in name_str:
        return "虎门局"
    elif "南区" in name_str:
        return "南区局"
    elif "西北" in name_str:
        return "西北局"
    elif "西区" in name_str:
        return "西区局"
    elif "长安" in name_str:
        return "长安局"
    return name_str


def find_column_indices(header_row):
    """Finds column indices for key fields in source files."""
    asset_col = None
    category_col = None
    district_col = None
    station_col = None
    time_col = None

    for col_idx, cell in enumerate(header_row, 1):
        val = str(cell.value).strip() if cell.value is not None else ""
        if not val:
            continue
        
        # 1. 资产编号
        if "资产" in val and "编号" in val:
            asset_col = col_idx
        # 2. 设备类别
        elif "设备" in val and "类别" in val:
            category_col = col_idx
        # 3. 县区局
        elif any(k in val for k in ["县区局", "区局", "供电局", "单位"]):
            district_col = col_idx
        # 4. 供电所
        elif "供电所" in val:
            station_col = col_idx
        # 5. 最后一次状态改变时间
        elif ("最后" in val and "状态" in val) or "改变时间" in val:
            time_col = col_idx

    return asset_col, category_col, district_col, station_col, time_col


# ==========================================
# Main Orchestration Function
# ==========================================
def main():
    # 1. Determine the base directory containing the running executable/script
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))

    # 2. Get current date to form the subdirectory path (YYYYMMDD)
    stats_time = datetime.datetime.now()
    date_str = stats_time.strftime("%Y%m%d")
    folder_path = os.path.join(base_dir, date_str)

    # Check if the subdirectory exists
    if not os.path.exists(folder_path):
        print(f"【错误】未找到当天的子目录：{folder_path}")
        print("请创建该目录，并将模板文件（待装提醒模板*.xlsm）和设备明细文件（*.xlsx）放入其中！")
        input("按任意键退出...")
        sys.exit(1)

    # 3. Check for template file matching '待装提醒模板*.xlsm' in the subdirectory
    template_pattern = os.path.join(folder_path, "待装提醒模板*.xlsm")
    templates = glob.glob(template_pattern)

    if not templates:
        print(f"【错误】未在子目录 '{folder_path}' 中找到 '待装提醒模板*.xlsm' 模板文件！")
        input("按任意键退出...")
        sys.exit(1)

    # Sort templates to pick the latest one alphabetically
    templates.sort()
    template_path = templates[-1]
    
    # Generate output file name in the same subdirectory
    output_filename = f"待装提醒{date_str}.xlsx"
    output_path = os.path.join(folder_path, output_filename)

    print(f"Working directory (Subdirectory): {folder_path}")
    print(f"Loading template: {template_path}")
    wb = openpyxl.load_workbook(template_path, keep_vba=False)

    # ------------------------------------------
    # Step 1 & 2: Merge Excel Files & Calculate Timeouts
    # ------------------------------------------
    ws_taizhang = wb["台帐信息"]
    
    # Clear old data in 台帐信息 (rows 2 onwards)
    if ws_taizhang.max_row >= 2:
        ws_taizhang.delete_rows(2, ws_taizhang.max_row)

    merged_rows = []

    # Find raw source files
    for filename in os.listdir(folder_path):
        if not filename.endswith(".xlsx"):
            continue
        # Skip output files or temp files
        if filename.startswith("待装提醒") or filename.startswith("~$"):
            continue

        file_full_path = os.path.join(folder_path, filename)
        print(f"Merging data from: {filename}")
        try:
            wb_src = openpyxl.load_workbook(file_full_path, data_only=True)
            for ws_src in wb_src.worksheets:
                # Find header row in first 10 rows
                header_row_idx = None
                header_cells = None
                
                rows_iter = ws_src.iter_rows(values_only=False)
                for r in range(1, 11):
                    try:
                        row_cells = next(rows_iter)
                    except StopIteration:
                        break
                    first_val = str(row_cells[0].value) if row_cells[0].value is not None else ""
                    if "资产编号" in first_val or "台帐信息" in first_val or "台账信息" in first_val:
                        header_row_idx = r
                        header_cells = row_cells
                        break
                
                if header_row_idx is None:
                    # Reset iterator and get first row
                    rows_iter = ws_src.iter_rows(values_only=False)
                    try:
                        header_cells = next(rows_iter)
                        header_row_idx = 1
                    except StopIteration:
                        continue

                # Map column indices
                cols = find_column_indices(header_cells)
                asset_col, category_col, district_col, station_col, time_col = cols

                if asset_col is None:
                    asset_col = 1

                # Read remaining rows from rows_iter
                for row_data in rows_iter:
                    asset_val = row_data[asset_col - 1].value
                    if asset_val is None or str(asset_val).strip() == "":
                        continue

                    # Extract basic fields
                    category_val = row_data[category_col - 1].value if category_col else ""
                    district_val = row_data[district_col - 1].value if district_col else ""
                    station_val = row_data[station_col - 1].value if station_col else ""
                    time_val = row_data[time_col - 1].value if time_col else None

                    # Handle state change time conversion
                    change_time = None
                    if isinstance(time_val, (datetime.datetime, datetime.date)):
                        change_time = time_val
                    elif isinstance(time_val, str):
                        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M", "%Y-%m-%d", "%Y/%m/%d"):
                            try:
                                change_time = datetime.datetime.strptime(time_val.strip(), fmt)
                                break
                            except ValueError:
                                pass

                    # Calculations
                    if change_time:
                        workday_dec = calculate_working_days_decimal(change_time, stats_time)
                        # Compare using the raw unrounded float workday_dec
                        over3 = "是" if workday_dec > 3.0 else "否"
                        over10 = "是" if workday_dec >= 10.0 else "否"
                        over15 = "是" if workday_dec >= 15.0 else "否"
                        last_month = is_last_month_received(change_time, stats_time)
                        time_str = change_time.strftime("%Y/%m/%d %H:%M")
                        # Round workday_dec to 2 decimal places for storage in Sheet
                        workday_dec_stored = round(workday_dec, 2)
                    else:
                        workday_dec_stored = "N/A"
                        over3 = "N/A"
                        over10 = "N/A"
                        over15 = "N/A"
                        last_month = "N/A"
                        time_str = "日期无效"

                    merged_rows.append([
                        asset_val,
                        category_val,
                        district_val,
                        station_val,
                        time_str,
                        stats_time.strftime("%Y/%m/%d %H:%M"),
                        workday_dec_stored,
                        over3,
                        over10,
                        over15,
                        last_month
                    ])
            wb_src.close()
        except Exception as e:
            print(f"Error processing file {filename}: {e}")

    # Write merged rows to 台帐信息 sheet
    for row_vals in merged_rows:
        ws_taizhang.append(row_vals)

    # Set formats/widths for 台帐信息
    for col in ws_taizhang.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_taizhang.column_dimensions[col_letter].width = max(max_len + 3, 12)

    print(f"Merged {len(merged_rows)} records into '台帐信息'.")

    # ------------------------------------------
    # Step 3 & 4: Aggregation and Notice Generation
    # ------------------------------------------
    ws_analysis = wb["区局分析"]
    
    # Find district rows to aggregate
    total_row_idx = None
    for r in range(2, 100):
        val = str(ws_analysis.cell(row=r, column=1).value).strip()
        if val in ["合计", "总计"]:
            total_row_idx = r
            break
    
    if total_row_idx is None:
        total_row_idx = 11

    # Extract original district rows (Row 2 to total_row_idx - 1)
    district_data = []
    for r in range(2, total_row_idx):
        dist_name = simplify_district_name(ws_analysis.cell(row=r, column=1).value)
        row_vals = [ws_analysis.cell(row=r, column=c).value for c in range(1, 11)]
        district_data.append({
            "name": dist_name,
            "original_row": row_vals,
            "counts": [0, 0, 0, 0, 0] # [待装台数, 上月已领出, 超3, 超10, 超15]
        })

    # Count from 台帐信息
    for row in merged_rows:
        raw_dist = row[2]
        if not raw_dist:
            continue
        simp_dist = simplify_district_name(raw_dist)

        # Find matching district
        for dist in district_data:
            if dist["name"] == simp_dist:
                dist["counts"][0] += 1 # Total
                if row[10] == "是":
                    dist["counts"][1] += 1 # Last month
                if row[7] == "是":
                    dist["counts"][2] += 1 # Over 3
                if row[8] == "是":
                    dist["counts"][3] += 1 # Over 10
                if row[9] == "是":
                    dist["counts"][4] += 1 # Over 15
                break

    # Sort the data rows based on Over 3 (counts[2]) descending
    district_data.sort(key=lambda x: x["counts"][2], reverse=True)

    # Write back sorted and calculated rows to 区局分析
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    no_fill = PatternFill(fill_type=None)

    for idx, dist in enumerate(district_data):
        r = 2 + idx
        ws_analysis.cell(row=r, column=1, value=dist["name"])
        ws_analysis.cell(row=r, column=2, value=dist["counts"][0]) # 待装数
        ws_analysis.cell(row=r, column=3, value=dist["counts"][1]) # 上月领出数
        ws_analysis.cell(row=r, column=4, value=dist["counts"][2]) # 超3个
        ws_analysis.cell(row=r, column=5, value=dist["counts"][3]) # 超10个
        ws_analysis.cell(row=r, column=6, value=dist["counts"][4]) # 超15个
        
        # Write back original G to J columns
        ws_analysis.cell(row=r, column=7, value=dist["original_row"][6])
        ws_analysis.cell(row=r, column=8, value=dist["original_row"][7])
        ws_analysis.cell(row=r, column=9, value=dist["original_row"][8])
        ws_analysis.cell(row=r, column=10, value=dist["original_row"][9])

        # Apply coloring
        for col_idx in (4, 5, 6):
            cell = ws_analysis.cell(row=r, column=col_idx)
            if isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.fill = yellow_fill
            else:
                cell.fill = no_fill

    # Calculate Totals
    total_total = sum(d["counts"][0] for d in district_data)
    total_last_month = sum(d["counts"][1] for d in district_data)
    total_over3 = sum(d["counts"][2] for d in district_data)
    total_over10 = sum(d["counts"][3] for d in district_data)
    total_over15 = sum(d["counts"][4] for d in district_data)

    # Write Totals
    ws_analysis.cell(row=total_row_idx, column=2, value=total_total)
    ws_analysis.cell(row=total_row_idx, column=3, value=total_last_month)
    ws_analysis.cell(row=total_row_idx, column=4, value=total_over3)
    ws_analysis.cell(row=total_row_idx, column=5, value=total_over10)
    ws_analysis.cell(row=total_row_idx, column=6, value=total_over15)

    # Get top 3 districts
    top1 = district_data[0]["name"] if len(district_data) > 0 else ""
    top2 = district_data[1]["name"] if len(district_data) > 1 else ""
    top3 = district_data[2]["name"] if len(district_data) > 2 else ""

    statement = (
        f"在计量设备待装超时监控方面，全市各区局累计已领取待装计量设备总数达 {total_total:,} 台。"
        f"其中，领出超 3 个工作日的计量设备 {total_over3:,} 台，"
        f"领出超 10 个工作日的计量设备 {total_over10:,} 台，"
        f"领出超 15 个工作日的计量设备 {total_over15:,} 台。"
        f"从超时情况来看，领出超时最多区局：{top1}、{top2}、{top3}。\n"
        f"    请有计量设备超时未安装的各相关区局负责人尽快把超时设备退库处理，避免影响计量设备安装不及时指标。"
    )
    ws_analysis.cell(row=14, column=1, value=statement)
    print("District analysis calculations and notice text written.")

    # ------------------------------------------
    # Step 5: Service Center Stats (供电服务中心)
    # ------------------------------------------
    ws_service = wb["供电服务中心"]
    ws_service.views.sheetView[0].showGridLines = True

    # Filter records where 县区局 is empty
    service_records = [r for r in merged_rows if not r[2]]

    # Initialize categories
    categories = {
        "电能表": {"count": 0, "last_month": 0, "over3": 0, "over10": 0, "over15": 0},
        "互感器": {"count": 0, "last_month": 0, "over3": 0, "over10": 0, "over15": 0},
        "负荷管理终端": {"count": 0, "last_month": 0, "over3": 0, "over10": 0, "over15": 0},
        "其他设备": {"count": 0, "last_month": 0, "over3": 0, "over10": 0, "over15": 0}
    }

    for row in service_records:
        cat_name = str(row[1])
        if "电能表" in cat_name or "电表" in cat_name:
            grp = "电能表"
        elif "互感器" in cat_name or "变压器" in cat_name:
            grp = "互感器"
        elif "终端" in cat_name or "负荷" in cat_name:
            grp = "负荷管理终端"
        else:
            grp = "其他设备"

        categories[grp]["count"] += 1
        if row[10] == "是":
            categories[grp]["last_month"] += 1
        if row[7] == "是":
            categories[grp]["over3"] += 1
        if row[8] == "是":
            categories[grp]["over10"] += 1
        if row[9] == "是":
            categories[grp]["over15"] += 1

    # Format table in 供电服务中心 sheet
    ws_service.delete_rows(1, ws_service.max_row + 10) # Clear all

    # Headers
    headers = ["设备类别", "台数", "上月已领出未安装", "领出超3工作日", "领出超10工作日", "领出超15工作日", "备注"]
    ws_service.append(headers)

    # Rows to write
    rows_to_write = []
    for cat in ["电能表", "互感器", "负荷管理终端"]:
        data = categories[cat]
        rows_to_write.append([cat, data["count"], data["last_month"], data["over3"], data["over10"], data["over15"], ""])
    if categories["其他设备"]["count"] > 0:
        data = categories["其他设备"]
        rows_to_write.append(["其他设备", data["count"], data["last_month"], data["over3"], data["over10"], data["over15"], ""])

    # Sum Totals
    tot_cnt = sum(categories[c]["count"] for c in categories)
    tot_lm = sum(categories[c]["last_month"] for c in categories)
    tot_o3 = sum(categories[c]["over3"] for c in categories)
    tot_o10 = sum(categories[c]["over10"] for c in categories)
    tot_o15 = sum(categories[c]["over15"] for c in categories)
    
    rows_to_write.append(["合计", tot_cnt, tot_lm, tot_o3, tot_o10, tot_o15, ""])

    for r in rows_to_write:
        ws_service.append(r)

    total_service_row = len(rows_to_write) + 1

    # Formatting service center sheet
    thin_border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    total_font = Font(name="Microsoft YaHei", size=10, bold=True)
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for r_idx in range(1, total_service_row + 1):
        ws_service.row_dimensions[r_idx].height = 22 if r_idx == 1 else 20
        for c_idx in range(1, 8):
            cell = ws_service.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif r_idx == total_service_row:
                cell.fill = total_fill
                cell.font = total_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if r_idx % 2 == 0:
                    cell.fill = zebra_fill
                if c_idx == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Microsoft YaHei", size=10)

    # Columns width
    for col in ws_service.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_service.column_dimensions[col_letter].width = max(max_len + 4, 15)

    print("Service center data aggregated and formatted.")

    # ------------------------------------------
    # Step 6: Export Clean XLSX Workbook
    # ------------------------------------------
    print(f"Saving final macro-free report to: {output_path}")
    wb.save(output_path)
    wb.close()
    
    # Copy output file to clipboard
    if copy_file_to_clipboard(output_path):
        print("【成功】已将生成的报表文件复制到系统剪贴板！可以直接在聊天窗口中按 Ctrl+V 粘贴发送。")
    else:
        print("【提示】未能将文件复制到剪贴板。")
        
    print("Done! Python script completed successfully.")


if __name__ == "__main__":
    main()
